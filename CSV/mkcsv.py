import sys
import re
import csv
import warnings
import tao.db
import re
from tao.util import unwrap as uw
from pprint import pformat
from datetime import datetime, time
from openpyxl import load_workbook
from MySQLdb.cursors import DictCursor
from pprint import pprint

class FileError(Exception):
  def __init__(self, fname):
    self.value = "[{}: ] couldn't match columns from".format(fname)
  def __str__(self):
    return self.value
  ## class FileError

class DateTimeError(Exception):
  def __init__(self, s, n, i):
    self.value = '[{}: {}] unknown value in column {}'.format(s, n, i)
  def __str__(self):
    return self.value
  ## class DateTimeError

class CharacterError(Exception):
  def __init__(self, s, n, e, r):
    self.value = pformat(dict(msg='[{}: {}] {}'.format(s, n, e), row=r))
  def __str__(self):
    return self.value
  ## class CharacterError

def date_column(s, N, c):
  val = None
  if c is None: 
    val = str()
  elif type(c) is datetime: 
    val = c.strftime('%Y-%m-%d')
  elif re.match(r"\?+", str(c).strip()):
    val = str()
  else: 
    raise DateTimeError(s, N, 0)
  return val
  ## def date_column

def time_column(s, N, c):
  val = None
  if c is None:
    val = str()
  else:
    if re.match(r"[0-9]{1,2}:[0-9]{2}-([0-9]{1,2}:[0-9]{2})", str(c).strip()):
      val = str(c).strip().split('-')[1]
    elif re.match(r"approx\.[ ]+[0-9]{1,2}:[0-9]{2}", str(c).strip()):
      val = str(c).strip().split()[1]
    elif re.match(r"[0-9]{1,2}:[0-9]{2}[ ]+GMT", str(c).strip()):
      val = str(c).strip().split()[0]
    elif re.match(r"[?]+", str(c).strip()):
      val = str()
    elif type(c) is time:
      val = c.strftime('%H:%M')
    elif type(c) is datetime:
      val = c.time().strftime('%H:%M')
    else:
      raise DateTimeError(s, N, 1)
  return val
  ## def time_column

def comment_column(c):
  val = c.replace(u"\u2019", "'")
  val = val.replace(u"\u2026", "...")
  return val
  ## def comment_column

def ocs_row_clean(row):
  csv_row = []
  for c in row:
    if c.value is None:
      continue
    elif type(c.value) is long:
      csv_row.append(str(c.value))
    elif re.search('\n',c.value) is not None:
      csv_row.append(c.value.replace("\n"," "))
    else:
      csv_row.append(c.value)
  return csv_row
  ## def ocs_row_clean

def ocs_inventory_sht(wb,sht):
  controllers = []
  faceplates = []
  for i,row in enumerate(wb[sht].iter_rows(min_row=2),2):
    if re.match(r"(TFLEX|FLEX)\s.+", row[0].value) is not None:
      sys.stdout.write('Controller Match: [{}: {}]\n'.format(sht, i))
      csv_row = ocs_row_clean(row)
      controllers.append(csv_row)
    elif re.match(r"Face Plate",row[0].value) is not None:
      sys.stdout.write('Face Plate Match: [{}: {}]\n'.format(sht, i))
      csv_row = ocs_row_clean(row)
      faceplates.append(csv_row)
  res = {'controllers': controllers, 'faceplates': faceplates}
  return res
  ## def ocs_inventory_sht

def make_history_csv(wb):
  for sht_name in [n for n in wb.sheetnames if n.isdigit()]:
    sht = wb[sht_name]
    row_iter = sht.rows
    N = 1
    ## skip first 3 rows
    for n in range(3): r = row_iter.next()
    N += n + 1
    ##
    with open('{}.csv'.format('TMBA'+sht_name), 'wb') as fob:
      csvobj = csv.writer(fob)
      for r in row_iter:
        csv_row = []
        ## column 0 (date)
        csv_row.append(date_column(sht_name, N, r[0].value))
        ## column 1 (time)
        csv_row.append(time_column(sht_name, N, r[1].value))
        ## column 2 (comment)
        if r[2].value is None:
          sys.stdout.write('[{}: {}] end of data\n'.format(sht_name, N))
          break
        else:
          csv_row.append(comment_column(r[2].value))
        try:
          csvobj.writerow(csv_row)
        except UnicodeEncodeError as uee:
          raise CharacterError(sht_name, N, uee, csv_row)
        N += 1
      ## end with file
  ## def make_history_csv

def make_iridium_csv(wb):
  sht_name = 'IridiumInfo'
  sht = wb[sht_name]
  row_iter = sht.rows
  N = 1
  ## skip first 3 rows
  for n in range(3): r = row_iter.next()
  N += n
  ##
  with open('TMBA{}.csv'.format(sht_name), 'wb') as fob:
    csvobj = csv.writer(fob)
    for r in row_iter:
      N += 1
      if r[0].value is None:
        sys.stdout.write('[{}: {}] end of data\n'.format(sht_name, N))
        break
      if any([c.value is None for c in r[2:5]]):
        sys.stdout.write('[{}: {}] missing SIM data\n'.format(sht_name, N))
        continue
      if any([not str(c.value).strip().isdigit() for c in r[0:5]]):
        sys.stdout.write('[{}: {}] bad number format\n'.format(sht_name, N))
        continue
      csv_row = [(((c.value is not None) and str(c.value).strip()) or str()) for c in r[0:6]]
      try:
        csvobj.writerow(csv_row)
      except UnicodeEncodeError as uee:
        raise CharacterError(sht_name, N, uee, csv_row)
  ## def make_iridium_csv

def make_current_hist(wb):
  sht = 'CurrentDrainTests'
  with open('TMBACurrentDrain.csv','wb') as fobj:
    csvobj = csv.writer(fobj)
    for i,row in enumerate(wb[sht].iter_rows(min_row=8)):
      if row[0] is not None:
        sn = row[0].value
        for c in row:
          csv_row = [sn]
          if c.value is not None and isinstance(c.value,datetime):
            csv_row.append(c.value)
            csv_row.append(wb[sht].cell(row=c.row, column=c.column+1).value)
            if c.fill.fgColor.rgb == 'FF00B050':
              # print('OceanServer: %s'%c.column,c.coordinate)
              csv_row.append('Test done with an OceanServer compass test')
            elif c.fill.fgColor.rgb == 'FF00B0F0':
              # print('Repair',c.coordinate)
              csv_row.append('Test done post repair')
            else:
              csv_row.append('')
            try:
              csvobj.writerow(csv_row)
            except UnicodeEncodeError as uee:
              raise CharacterError(active, i, uee, csv_row)

def make_status_csv(wb):
  sht_name = 'STATUS'
  sht = wb[sht_name]
  row_iter = sht.rows
  N = 1
  ## skip first 3 rows
  for n in range(3): row_iter.next()
  N += n + 1
  ##
  with open('TMBA{}.csv'.format(sht_name), 'wb') as fob:
    csvobj = csv.writer(fob)
    for r in row_iter:
      if r[0].value is None:
        sys.stdout.write('[{}: {}] end of data\n'.format(sht_name, N))
        break
      csv_row = [(((c.value is not None) and str(c.value).strip()) or str()) for c in r]
      try:
        csvobj.writerow(csv_row)
      except UnicodeEncodeError as uee:
        raise CharacterError(sht_name, N, uee, csv_row)
      N += 1
  ## def make_status_csv

def fetch_ocs_status():
  dbc = tao.db.MySQL().user()
  crs = dbc.cursor(DictCursor)
  sql = r"""
    SELECT SUBSTR(`sn`,-4) AS 'sn', MIN(`dt`) as 'dt' FROM `Flex`.`FlexHist`
    WHERE `sn` LIKE '%FLEX%' GROUP BY `sn`
    """
  n = crs.execute(uw(sql))
  d = crs.fetchall()
  dbc.close()
  return d
  ## def fetch_info_data

def fetch_ocs_history(sn):
  dbc = tao.db.MySQL().user()
  crs = dbc.cursor(DictCursor)
  sql = r"""
    SELECT `dt`,`comment` FROM `Flex`.`FlexHist`
    WHERE `sn` LIKE '%FLEX%{0}%'
    """
  n = crs.execute(uw(sql.format(sn)))
  d = crs.fetchall()
  dbc.close()
  return d
  ## def fetch_ocs_history

def make_ocs_csv(wb):
  sht_name = 'OCS'
  sht = wb[sht_name]
  row_iter = sht.rows
  N = 1
  ## skip first 1 rows
  for n in range(1): row_iter.next()
  N += n + 1
  ##
  with open('{}Iridium.csv'.format(sht_name), 'wb') as fob:
    csvobj = csv.writer(fob)
    for r in row_iter:
      csv_row = [(((c.value is not None) and str(c.value).strip()) or str()) for c in r]
      try:
        csvobj.writerow(csv_row)
      except UnicodeEncodeError as uee:
        raise CharacterError(sht_name, N, uee, csv_row)
      N += 1
  ## def make_ocs_csv

def make_ocs_hist(sns):
  for sn in sns:
    hist = fetch_ocs_history(sn)
    box = r'0\d{3}'
    sn = 'BOX'+sn if re.match(box,sn) is not None else sn
    # for ent in hist:
    #   ent['comment'] = comment_column(ent['comment'])
    with open('OCS{}.csv'.format(sn),'wb') as fobj:
      header = ['dt','comment']
      csvobj = csv.DictWriter(fobj,fieldnames=header)
      csvobj.writerows(hist)
  ## def make_ocs_hist

def make_ocs_inv(wb):
  active = 'In Service'
  retire = 'Retired'
  with open('OCSInventory.csv', 'w') as fobj:
    csvobj = csv.writer(fobj)
    dbstat = fetch_ocs_status()
    make_ocs_hist([r['sn'] for r in dbstat])
    box = r'0\d{3}'
    for r in dbstat:
      if re.match(box,r['sn']):
        r['sn'] = 'BOX'+r['sn']
    db = [[r['sn'],r['dt']] for r in dbstat]
    act = ocs_inventory_sht(wb,active)
    ret = ocs_inventory_sht(wb,retire)
    inv = []
    inv.extend(db+act['controllers']+ret['controllers'])
    plates = []
    plates.extend(act['faceplates']+ret['faceplates'])
    for ent in inv:
      try:
        csvobj.writerow(ent)
      except UnicodeEncodeError as uee:
        raise CharacterError(active, i, uee, ent)
    
  ## def make_ocs_inv

def make_stratos_csv(wb):
  sht_name = 'Mobile Info & Summary'
  sht = wb[sht_name]
  row_iter = sht.rows
  N = 1
  ## skip first 1 rows
  for n in range(4): row_iter.next()
  N += n
  ##
  with open('{}.csv'.format('Stratos'), 'wb') as fob:
    csvobj = csv.writer(fob)
    for r in row_iter:
      N += 1
      if r[0].value is None:
        sys.stdout.write('[{}: {}] end of data\n'.format(sht_name, N))
        break
      if r[8].value is None or not str(r[8].value).strip():
        ## skip the odd entries
        sys.stdout.write('[{}: {}] missing SIM number\n'.format(sht_name, N))
        continue
      csv_row = [str(int(r[8].value)), None, None]
      csv_row[1] = ((r[2].value is not None) and str(r[2].value).strip()) or str()
      csv_row[2] = ((r[3].value is not None) and str(r[3].value).strip()) or str()
      try:
        csvobj.writerow(csv_row)
      except UnicodeEncodeError as uee:
        raise CharacterError(sht_name, N, uee, csv_row)
  ## def make_stratos_csv

def main(args):
  for xlsxfile in args:
    sys.stdout.write('Processing: {}\n'.format(xlsxfile))
    with warnings.catch_warnings(record=True) as wrn:
      wb = load_workbook(xlsxfile, read_only=True, data_only=True)
      if wrn:
        sys.stdout.write('[{}]\n\topenpyxl: {}\n'.format(xlsxfile, wrn.pop().message))
    ##
    try:
      if set(['TFLEX','OCS']).issubset(wb.sheetnames):
        make_ocs_csv(wb)
      elif set(['IridiumInfo','STATUS']).issubset(wb.sheetnames):
        make_history_csv(wb)
        make_current_hist(wb)
        make_iridium_csv(wb)
        make_status_csv(wb)
      elif 'Mobile Info & Summary' in wb.sheetnames:
        make_stratos_csv(wb)
      elif set(['In Service','Retired']).issubset(wb.sheetnames):
        make_ocs_inv(wb)
      else:
        raise FileError(xlsxfile)
    except FileError as err:
      return str(err)
  ## def main

if __name__ == '__main__':
  sys.exit(main(sys.argv[1:]))
